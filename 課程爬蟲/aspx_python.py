import time
import os
import pandas as pd
from io import StringIO
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
import random
import logging
from openpyxl import load_workbook

# --- 導入 selenium-stealth ---
from selenium_stealth import stealth

# --- 設定日誌 ---
logging.basicConfig(filename='course_scraper.log', level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    encoding='utf-8')

# --- 步驟 1: 分析與設定 ---
# 更新：將入口網址改為首頁/公告頁
ENTRY_URL = "https://stu.mcu.edu.tw/appx/qry/Qs01.aspx"

# --- 步驟 2: 使用 Selenium 模擬瀏覽器操作並獲取課表 ---
def fetch_course_data_with_selenium(driver, department_id, grade_id):
    """
    在已經進入查詢頁面的前提下，選擇系所與年級並抓取資料。
    """
    try:
        logging.info(f"正在查詢 [系所:{department_id}, 年級:{grade_id}]")
        print(f"正在查詢 [系所:{department_id}, 年級:{grade_id}] 的課表...")

        wait = WebDriverWait(driver, 20)
        
        # 1. 等待並選擇系所
        dept_dropdown_element = wait.until(
            EC.element_to_be_clickable((By.ID, 'MainContent_drpDepartment'))
        )
        Select(dept_dropdown_element).select_by_value(department_id)

        # 2. 等待並選擇年級
        grade_dropdown_element = wait.until(
            EC.element_to_be_clickable((By.ID, 'MainContent_drpGrade'))
        )
        Select(grade_dropdown_element).select_by_value(grade_id)

        # 3. 點擊「送出」按鈕
        submit_button = wait.until(
            EC.element_to_be_clickable((By.ID, 'MainContent_btnSend'))
        )
        submit_button.click()

        # 4. 等待課表表格出現
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#printlist table")))
        
        page_source = driver.page_source
        df_list = pd.read_html(StringIO(page_source))
        
        course_df = df_list[-1]
        course_df.dropna(how='all', inplace=True)

        if not course_df.empty:
            course_df['系所代碼'] = department_id
            course_df['年級'] = grade_id
            logging.info(f"成功解析到 {len(course_df)} 筆課程資料")
            print(f"  > 成功解析到 {len(course_df)} 筆課程資料")
            return course_df
        else:
            logging.warning(f"  > 查無資料或表格為空。")
            print(f"  > 查無資料或表格為空。")
            return pd.DataFrame()

    except TimeoutException:
        logging.warning(f"  > 查無資料 (等待超時)。")
        print(f"  > 查無資料 (等待超時)。")
        return pd.DataFrame()
    except Exception as e:
        logging.error(f"  > 查詢或解析時發生錯誤: {e}")
        print(f"  > 查詢或解析時發生錯誤: {e}")
        return pd.DataFrame()


# --- 步驟 3: 迭代所有系所與年級 ---
def main():
    GRADES_TO_FETCH = ['1', '2', '3', '4']
    
    department_list = {
        "01": "英語教學中心", "02": "體育室", "09": "數位媒體設計學系", "11": "企業管理學系",
        "13": "資訊管理學系", "14": "觀光事業學系", "15": "影音新聞暨社群傳播學系",
        "17": "應用統計與資料科學學系", "18": "休閒遊憩管理學系", "19": "餐旅管理學系",
        "21": "商業設計學系", "23": "商品設計學系", "24": "建築學系", "26": "廣告暨策略行銷學系",
        "27": "新聞與大眾傳播學士學位學程", "30": "傳播學院", "31": "新媒體暨傳播管理學系",
        "32": "廣播電視學系", "36": "資訊工程學系", "38": "犯罪防治學系", "39": "生物科技學系",
        "40": "醫療資訊與管理學系", "41": "法律學系", "42": "應用英語學系", "44": "應用日語學系",
        "46": "金融科技應用學士學位學程", "48": "旅遊與觀光學士學位學程", "52": "會計學系",
        "53": "資訊應用與金融保險學系", "54": "財務金融學系", "57": "國際企業學系",
        "59": "國際事務與外交學士學位學程", "60": "諮商臨床與工商心理學系",
        "61": "諮商臨床與工商心理學系碩士班諮商心理學組", "62": "諮商臨床與工商心理學系碩士班臨床心理學組",
        "64": "金融學系", "65": "財金法律學系", "67": "生物醫學工程學系", "68": "公共事務與行政管理學系",
        "70": "高階經理碩士在職學位學程", "74": "都市設計與永續發展學系", "75": "人工智慧應用學系",
        "76": "半導體應用學士學位學程", "79": "應用中文與華語文教學系", "83": "教育研究所", "84": "電機工程學系",
        "86": "金融科技應用學系", "88": "時尚與創新管理學士學位學程",
        "89": "資訊科技應用與管理學士學位學程", "90": "國際事務碩士學位學程",
        "91": "國際企業與貿易學士學位學程", "96": "金融科技碩士學位學程",
        "98": "國防行政管理進修學士學位學程", "99": "國際班", "KM": "金門校區(不分系所)"
    }
    
    output_filename = '全校課表_Selenium.xlsx'
    
    if not os.path.exists(output_filename):
        pd.DataFrame().to_excel(output_filename, index=False)

    for dept_id, dept_name in department_list.items():
        driver = None
        try:
            chrome_options = Options()
            chrome_options.add_argument("--headless")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("start-maximized")
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            
            # --- 更新：使用 try-except 處理不同 Selenium 版本的初始化語法 ---
            try:
                # 新版 Selenium (4.x) 語法
                service = Service(ChromeDriverManager().install())
                driver = webdriver.Chrome(service=service, options=chrome_options)
            except TypeError:
                # 舊版 Selenium (3.x) 語法
                logging.warning("偵測到舊版 Selenium 語法，切換至 executable_path。")
                driver = webdriver.Chrome(executable_path=ChromeDriverManager().install(), options=chrome_options)


            stealth(driver, languages=["en-US", "en"], vendor="Google Inc.", platform="Win32",
                    webgl_vendor="Intel Inc.", renderer="Intel Iris OpenGL Engine", fix_hairline=True)

            # --- 更新：模擬真人操作流程 ---
            # 1. 訪問入口頁面
            print(f"正在處理系所: {dept_name} ({dept_id})")
            driver.get(ENTRY_URL)
            
            # 2. 點擊「必修課程」連結，進入查詢頁面
            wait = WebDriverWait(driver, 15)
            required_courses_link = wait.until(
                EC.element_to_be_clickable((By.LINK_TEXT, "必修課程"))
            )
            required_courses_link.click()
            print("  > 已成功點擊「必修課程」，進入查詢頁面。")
            
            dept_dfs = []
            for grade in GRADES_TO_FETCH:
                df = fetch_course_data_with_selenium(driver, dept_id, grade)
                if not df.empty:
                    df['系所名稱'] = dept_name
                    dept_dfs.append(df)
                time.sleep(random.uniform(0.5, 1.5))

            if dept_dfs:
                combined_df = pd.concat(dept_dfs, ignore_index=True)
                book = load_workbook(output_filename)
                with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    is_empty = writer.sheets['Sheet1'].max_row <= 1
                    combined_df.to_excel(writer, index=False, header=is_empty, startrow=writer.sheets['Sheet1'].max_row if not is_empty else 0)
                print(f"  > 已將系所 {dept_name} 的資料寫入 Excel。")

        except Exception as e:
            logging.error(f"處理系所 {dept_name} ({dept_id}) 時發生嚴重錯誤: {e}")
            print(f"處理系所 {dept_name} ({dept_id}) 時發生嚴重錯誤: {e}")
        finally:
            if driver:
                driver.quit()
        
        logging.info(f"完成系所 {dept_name} ({dept_id}) 的處理")
        print("-" * 50)

    print(f"\n爬取完成！資料已儲存至 {os.path.abspath(output_filename)}")


if __name__ == "__main__":
    main()
